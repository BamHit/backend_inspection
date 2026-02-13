"""Générateur de rapports Excel pour l'historique des inspections."""

import base64
import logging
from io import BytesIO
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

logger = logging.getLogger(__name__)


class ExcelHistoryBuilder:
    """Construit un fichier Excel avec l'historique des inspections."""

    # Styles
    HEADER_FILL = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ANOMALY_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Colonnes
    HEADERS = [
        "Timestamp",
        "Part Number",
        "Face",
        "Result",
        "ROI Index",
        "Expected Class",
        "Predicted Class",
        "Probability",
        "ROI Coordinates",
        "Reference ROI Image",
        "Inspected ROI Image",
        "Reference Face with Overlay",
        "Inspected Face with Overlay",
    ]

    COLUMN_WIDTHS = {
        "A": 30,  # Timestamp
        "B": 15,  # Part Number
        "C": 12,  # Face
        "D": 12,  # Result
        "E": 12,  # ROI Index
        "F": 20,  # Expected Class
        "G": 20,  # Predicted Class
        "H": 12,  # Probability
        "I": 20,  # ROI Coordinates
        "J": 25,  # Reference ROI Image
        "K": 25,  # Inspected ROI Image
        "L": 30,  # Reference Face Overlay
        "M": 30,  # Inspected Face Overlay
    }

    def __init__(self) -> None:
        """Initialise le builder."""
        self.wb: Workbook | None = None
        self.ws: Worksheet | None = None

    def build(self, history_data: list[dict[str, Any]]) -> Workbook:
        """Construit le workbook Excel complet.

        Args:
            history_data: Liste des inspections à exporter.

        Returns:
            Workbook Excel prêt à être sauvegardé.
        """
        logger.info(f"Building Excel report for {len(history_data)} inspections")

        self.wb = Workbook()
        # On force le type Worksheet car wb.active peut être une ReadOnlyWorksheet
        self.ws = cast(Worksheet, self.wb.active)
        self.ws.title = "Inspection History"

        self._create_header()
        self._set_column_widths()

        row_num = 2
        for inspection in history_data:
            row_num = self._add_inspection(inspection, row_num)

        logger.info("Excel report built successfully")
        return self.wb

    def _create_header(self) -> None:
        """Crée la ligne d'en-tête."""
        assert self.ws is not None
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

    def _set_column_widths(self) -> None:
        """Définit les largeurs de colonnes."""
        assert self.ws is not None
        for col_letter, width in self.COLUMN_WIDTHS.items():
            self.ws.column_dimensions[col_letter].width = width

    def _add_inspection(self, inspection: dict[str, Any], row_num: int) -> int:
        """Ajoute une inspection complète au workbook."""
        timestamp = inspection.get("timestamp", "N/A")
        part_number = inspection.get("partNumber", "N/A")
        faces = inspection.get("faces", [])

        for face in faces:
            row_num = self._add_face(timestamp, part_number, face, row_num)

        return row_num

    def _add_face(
        self, timestamp: str, part_number: str, face: dict[str, Any], row_num: int
    ) -> int:
        """Ajoute une face (avec ses ROIs) au workbook."""
        face_name = face.get("faceName", "N/A")
        rois = face.get("rois", [])
        has_anomaly = face.get("hasAnomaly", False)
        result_str = "REJECT" if has_anomaly else "PASS"

        # Charger les images de la face
        ref_img_pil, test_img_pil = self._load_face_images(face)
        ref_overlay_xl, test_overlay_xl = self._create_overlays(ref_img_pil, test_img_pil, rois)

        first_row = row_num

        # Ajouter chaque ROI
        for roi_idx, roi in enumerate(rois):
            row_num = self._add_roi(
                timestamp,
                part_number,
                face_name,
                result_str,
                roi_idx,
                roi,
                ref_img_pil,
                test_img_pil,
                row_num,
            )

        # Ajouter les overlays (fusionnés sur toutes les ROIs de cette face)
        self._add_face_overlays(ref_overlay_xl, test_overlay_xl, first_row, row_num)

        return row_num

    def _load_face_images(
        self, face: dict[str, Any]
    ) -> tuple[PILImage.Image | None, PILImage.Image | None]:
        """Charge les images de référence et de test."""
        ref_img_pil = None
        test_img_pil = None

        ref_image_b64 = face.get("refImage")
        if ref_image_b64:
            try:
                img_data = base64.b64decode(
                    ref_image_b64.split(",")[1] if "," in ref_image_b64 else ref_image_b64
                )
                ref_img_pil = PILImage.open(BytesIO(img_data))
            except Exception as e:
                logger.warning(f"Error loading reference image: {e}")

        test_image_b64 = face.get("testImage")
        if test_image_b64:
            try:
                img_data = base64.b64decode(
                    test_image_b64.split(",")[1] if "," in test_image_b64 else test_image_b64
                )
                test_img_pil = PILImage.open(BytesIO(img_data))
            except Exception as e:
                logger.warning(f"Error loading test image: {e}")

        return ref_img_pil, test_img_pil

    def _create_overlays(
        self,
        ref_img: PILImage.Image | None,
        test_img: PILImage.Image | None,
        rois: list[dict[str, Any]],
    ) -> tuple[XLImage | None, XLImage | None]:
        """Crée les images overlay avec ROIs dessinées."""
        ref_overlay = None
        test_overlay = None

        if ref_img:
            try:
                overlay_img = self._draw_rois_on_image(ref_img, rois)
                img_bytes, w, h = self._compress_image(overlay_img, max_width=1000, quality=85)
                ref_overlay = XLImage(img_bytes)
                ref_overlay.width = 250
                ref_overlay.height = int(250 * h / w)
            except Exception as e:
                logger.warning(f"Error creating reference overlay: {e}")

        if test_img:
            try:
                overlay_img = self._draw_rois_on_image(test_img, rois)
                img_bytes, w, h = self._compress_image(overlay_img, max_width=1000, quality=85)
                test_overlay = XLImage(img_bytes)
                test_overlay.width = 250
                test_overlay.height = int(250 * h / w)
            except Exception as e:
                logger.warning(f"Error creating test overlay: {e}")

        return ref_overlay, test_overlay

    def _add_roi(
        self,
        timestamp: str,
        part_number: str,
        face_name: str,
        result_str: str,
        roi_idx: int,
        roi: dict[str, Any],
        ref_img: PILImage.Image | None,
        test_img: PILImage.Image | None,
        row_num: int,
    ) -> int:
        """Ajoute une ligne de ROI au worksheet."""
        assert self.ws is not None

        expected_class = roi.get("expectedClass", "N/A")
        predicted_class = roi.get("predicted", "N/A")
        probability = float(roi.get("probability", 0.0))
        is_anomaly = roi.get("isAnomaly", False)

        x, y, size = roi.get("x", 0), roi.get("y", 0), roi.get("size", 0)
        roi_coords = f"({int(x)}, {int(y)}, {int(size)})"

        self.ws.cell(row=row_num, column=1, value=timestamp).border = self.BORDER
        self.ws.cell(row=row_num, column=2, value=part_number).border = self.BORDER
        self.ws.cell(row=row_num, column=3, value=face_name).border = self.BORDER
        self.ws.cell(row=row_num, column=4, value=result_str).border = self.BORDER
        self.ws.cell(row=row_num, column=5, value=f"ROI {roi_idx + 1}").border = self.BORDER
        self.ws.cell(row=row_num, column=6, value=expected_class).border = self.BORDER
        self.ws.cell(row=row_num, column=7, value=predicted_class).border = self.BORDER
        self.ws.cell(row=row_num, column=8, value=f"{probability:.2%}").border = self.BORDER
        self.ws.cell(row=row_num, column=9, value=roi_coords).border = self.BORDER

        if is_anomaly:
            for col in range(1, 10):
                self.ws.cell(row=row_num, column=col).fill = self.ANOMALY_FILL

        self._add_roi_images(ref_img, test_img, roi, row_num)
        return row_num + 1

    def _add_roi_images(
        self,
        ref_img: PILImage.Image | None,
        test_img: PILImage.Image | None,
        roi: dict[str, Any],
        row_num: int,
    ) -> None:
        """Ajoute les images croppées des ROIs."""
        assert self.ws is not None

        if ref_img:
            try:
                roi_cropped = self._crop_roi(ref_img, roi)
                img_bytes, w, h = self._compress_image(roi_cropped, max_width=150, quality=60)
                xl_img = XLImage(img_bytes)
                xl_img.width = 100
                xl_img.height = int(100 * h / w)
                self.ws.add_image(xl_img, f"J{row_num}")
                current_height = float(self.ws.row_dimensions[row_num].height or 15)
                self.ws.row_dimensions[row_num].height = max(current_height, xl_img.height * 0.75)
            except Exception as e:
                logger.warning(f"Error adding reference ROI image: {e}")

        if test_img:
            try:
                roi_cropped = self._crop_roi(test_img, roi)
                img_bytes, w, h = self._compress_image(roi_cropped, max_width=150, quality=60)
                xl_img = XLImage(img_bytes)
                xl_img.width = 100
                xl_img.height = int(100 * h / w)
                self.ws.add_image(xl_img, f"K{row_num}")
                current_height = float(self.ws.row_dimensions[row_num].height or 15)
                self.ws.row_dimensions[row_num].height = max(current_height, xl_img.height * 0.75)
            except Exception as e:
                logger.warning(f"Error adding test ROI image: {e}")

    def _add_face_overlays(
        self,
        ref_overlay: XLImage | None,
        test_overlay: XLImage | None,
        first_row: int,
        last_row: int,
    ) -> None:
        """Ajoute les overlays face complète."""
        assert self.ws is not None

        if ref_overlay and first_row < last_row:
            if last_row - first_row > 1:
                self.ws.merge_cells(f"L{first_row}:L{last_row - 1}")
            self.ws.add_image(ref_overlay, f"L{first_row}")

        if test_overlay and first_row < last_row:
            if last_row - first_row > 1:
                self.ws.merge_cells(f"M{first_row}:M{last_row - 1}")
            self.ws.add_image(test_overlay, f"M{first_row}")

    @staticmethod
    def _draw_rois_on_image(img: PILImage.Image, rois: list[dict[str, Any]]) -> PILImage.Image:
        """Dessine les ROIs sur une image PIL."""
        import cv2
        import numpy as np

        img_cv = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)

        for idx, roi in enumerate(rois):
            x = int(roi.get("x", 0))
            y = int(roi.get("y", 0))
            size = int(roi.get("size", 0))
            expected_class = roi.get("expectedClass", "ROI")

            cv2.rectangle(img_cv, (x, y), (x + size, y + size), (2, 74, 216), 3)
            label = f"ROI {idx + 1} - {expected_class}"
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 0.5
            thickness = 2
            (text_w, text_h), _ = cv2.getTextSize(label, font, font_scale, thickness)

            cv2.rectangle(img_cv, (x, y - text_h - 10), (x + text_w + 10, y), (2, 74, 216), -1)
            cv2.putText(img_cv, label, (x + 5, y - 5), font, font_scale, (255, 255, 255), thickness)

        return PILImage.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))

    @staticmethod
    def _crop_roi(img: PILImage.Image, roi: dict[str, Any]) -> PILImage.Image:
        """Croppe une ROI depuis une image."""
        x = int(roi.get("x", 0))
        y = int(roi.get("y", 0))
        size = int(roi.get("size", 0))

        margin = 5
        left = max(0, x - margin)
        top = max(0, y - margin)
        right = min(img.width, x + size + margin)
        bottom = min(img.height, y + size + margin)

        return img.crop((left, top, right, bottom))

    @staticmethod
    def _compress_image(
        img: PILImage.Image, max_width: int = 300, quality: int = 60
    ) -> tuple[BytesIO, int, int]:
        """Compresse une image PIL pour Excel."""
        if img.width > max_width:
            ratio = max_width / img.width
            new_size = (max_width, int(img.height * ratio))
            img = img.resize(new_size, PILImage.Resampling.LANCZOS)

        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format="JPEG", quality=quality, optimize=True)
        img_byte_arr.seek(0)

        return img_byte_arr, img.width, img.height
